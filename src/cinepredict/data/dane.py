"""Procesa el anexo de proyecciones del DANE (municipal por edad) a features.

Fuente: PPED-AreaSexoEdadMun-2018-2042_VP.xlsx (DANE, base Censo 2018).
Hoja `PobMunicipalxÁreaSexoEdad`: por municipio × año × área geográfica trae la
población por sexo y edades simples (0 a 100). Tomamos el área "Total" y sumamos
la franja **15–44 años**, que concentra el mayor consumo de cine, para alimentar
el Componente A (demanda potencial municipal).

Genera `data/processed/dane_poblacion.parquet` con, por municipio y año:
  cod_divipola, anio, poblacion_total, poblacion_15_44, prop_15_44
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pandas as pd
from loguru import logger

from cinepredict.config import EXTERNAL_DIR, PROCESSED_DIR

SHEET = "PobMunicipalxÁreaSexoEdad"
AREA_OBJETIVO = "Total"          # excluye Cabecera / Centros Poblados para no duplicar
HEADER_ROW = 9                   # fila con los nombres detallados de columnas
EDAD_MIN, EDAD_MAX = 15, 44      # franja de mayor consumo cinematográfico

# Índices posicionales (1-based) de las primeras columnas (encabezado combinado)
COL_MPIO = 3      # código DIVIPOLA de 5 dígitos
COL_ANIO = 5
COL_AREA = 6
COL_TOTAL = 7     # población total (ambos sexos, todas las edades)

DEFAULT_XLSX = EXTERNAL_DIR / "dane_proyecciones_municipal.xlsx"


def _detectar_columnas_edad(header: tuple) -> list[int]:
    """Devuelve los índices (0-based) de las columnas 'Total <edad> años' en 15–44."""
    objetivo = []
    for idx, name in enumerate(header):
        if not isinstance(name, str):
            continue
        m = re.match(r"^Total (\d+)", name.strip())
        if m and EDAD_MIN <= int(m.group(1)) <= EDAD_MAX:
            objetivo.append(idx)
    return objetivo


def build_dane_features(xlsx_path: Path = DEFAULT_XLSX) -> Path:
    """Lee el anexo del DANE en streaming y construye la tabla de features."""
    if not xlsx_path.exists():
        raise FileNotFoundError(
            f"No existe {xlsx_path}. Descárgalo con: cinepredict download --source dane"
        )

    logger.info(f"Abriendo {xlsx_path.name} (modo solo-lectura)…")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb[wb.sheetnames[-1]]

    cols_edad: list[int] = []
    registros: list[dict] = []

    for n, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if n == HEADER_ROW:
            cols_edad = _detectar_columnas_edad(row)
            logger.info(f"Detectadas {len(cols_edad)} columnas de edad (15–44).")
            continue
        if n <= HEADER_ROW or not row:
            continue
        if row[COL_AREA - 1] != AREA_OBJETIVO:
            continue  # solo el agregado 'Total' del municipio

        mpio = row[COL_MPIO - 1]
        if not mpio:
            continue
        pob_total = row[COL_TOTAL - 1] or 0
        pob_15_44 = sum((row[i] or 0) for i in cols_edad)
        registros.append({
            "cod_divipola": str(mpio).zfill(5),
            "anio": int(row[COL_ANIO - 1]),
            "poblacion_total": int(pob_total),
            "poblacion_15_44": int(pob_15_44),
        })

    wb.close()

    df = pd.DataFrame.from_records(registros)
    df["prop_15_44"] = (df["poblacion_15_44"] / df["poblacion_total"]).round(4)

    out = PROCESSED_DIR / "dane_poblacion.parquet"
    out.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(out, index=False)
    logger.success(
        f"DANE features: {len(df):,} filas "
        f"({df['cod_divipola'].nunique()} municipios × {df['anio'].nunique()} años) -> {out}"
    )
    return out
